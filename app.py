import os
import pickle
# Gmail API utils
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from google.oauth2 import service_account
# for encoding/decoding messages in base64
from base64 import urlsafe_b64decode, urlsafe_b64encode
# for dealing with attachement MIME types
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
from email.mime.audio import MIMEAudio
from email.mime.base import MIMEBase
from email import encoders
from mimetypes import guess_type as guess_mime_type

import openpyxl 
import time

# Request all access (permission to read/send/receive emails, manage the inbox, and more)
#SCOPES = ['https://mail.google.com/']
our_email = 'mhadin96@gmail.com'
SCOPES = ['https://www.googleapis.com/auth/gmail.send']


# credentials = service_account.Credentials.from_service_account_file(
#     'client_secret.json', scopes=SCOPES)


def gmail_authenticate():
    creds = None
    # the file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first time
    if os.path.exists("token.pickle"):
        with open("token.pickle", "rb") as token:
            creds = pickle.load(token)
    # if there are no (valid) credentials availablle, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file('client_secret.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # save the credentials for the next run
        with open("token.pickle", "wb") as token:
            pickle.dump(creds, token)
    return build('gmail', 'v1', credentials=creds)


def add_attachment(message, filename):
    content_type, encoding = guess_mime_type(filename)
    if content_type is None or encoding is not None:
        content_type = 'application/octet-stream'
    main_type, sub_type = content_type.split('/', 1)
    if main_type == 'text':
        fp = open(filename, 'rb')
        msg = MIMEText(fp.read().decode(), _subtype=sub_type)
        fp.close()
    elif main_type == 'image':
        fp = open(filename, 'rb')
        msg = MIMEImage(fp.read(), _subtype=sub_type)
        fp.close()
    elif main_type == 'audio':
        fp = open(filename, 'rb')
        msg = MIMEAudio(fp.read(), _subtype=sub_type)
        fp.close()
    else:
        fp = open(filename, 'rb')
        msg = MIMEBase('application', 'octet-stream')
        msg.set_payload(fp.read())
        encoders.encode_base64(msg)
        fp.close()
    filename = os.path.basename(filename)
    msg.add_header('Content-Disposition', 'attachment', filename=filename)
    message.attach(msg)



def build_message(destination, obj, body, attachments=[]):
    if not attachments: # no attachments given
        message = MIMEText(body,'html')
        message['to'] = destination
        message['from'] = our_email
        message['subject'] = obj
    else:
        message = MIMEMultipart()
        message['to'] = destination
        message['from'] = our_email
        message['subject'] = obj
        message.attach(MIMEText(body, 'html'))
        for filename in attachments:
            add_attachment(message, filename)
    return {'raw': urlsafe_b64encode(message.as_bytes()).decode()}
                




def send_message(service, destination, obj, body, attachments=[]):
    return service.users().messages().send(
      userId="me",
      body=build_message(destination, obj, body, attachments)
    ).execute()



text1="""<td width="708" valign="top" style="width:531pt;border:1pt solid windowtext;padding:0in 5.4pt">
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif"><p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">&nbsp;</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">Dear
  Professor """ #Last Name

text2=""",</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">&nbsp;</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">I hope this
  email finds you well. My name is MohammadHadi Najafi, and I am writing to
  express my strong interest in pursuing a Direct PhD under your supervision in
  the """ #Department
text3=" at " #University
text4=". I am particularly drawn to your groundbreaking work in " #Interest
text5=""",
  and I believe my background aligns well with your research interests.</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">&nbsp;</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">I have
  completed my Bachelor’s degree in Electrical Engineering at Quchan University
  of Technology. Throughout my academic journey, I have developed a deep
  passion for robotics and embedded systems design. My Bachelor’s thesis
  focuses on “Software-Based TFT LCD Driver for ARM Microcontrollers without
  LTDC”, which has prepared me well for advanced development of embedded
  systems.</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">&nbsp;</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">What sets me
  apart is my track record of success in both academic competitions and
  industry experience:</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif"><b>1.</b> <b>Robotics
  Competitions</b>: I have achieved significant success in several national and
  international robotics competitions. </p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif"><b>2.</b> <b>Professional
  Experience</b>: I have 5 years of hands-on experience in electronics and
  embedded system design. </p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif"><b>3.</b> <b>Technical
  Skills</b>: I possess strong skills in C, Python, MATLAB and LabView
  programming, Embedded Linux, PCB design, IOT and sensor integration. </p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">&nbsp;</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">I believe my
  combination of theoretical knowledge, practical experience, and
  problem-solving skills would allow me to contribute significantly to this
  line of research and expand upon it. I would be grateful for the opportunity
  to discuss potential PhD positions in your lab and how my background might
  complement your current research initiatives. I have attached my CV, which
  provides more details about my projects and achievements. If you need any
  additional information or work samples, please don't hesitate to ask.</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">&nbsp;</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">Thank you for
  your time and consideration. I am excited about the possibility of
  contributing to the cutting-edge research in your lab and look forward to the
  opportunity to discuss this further.</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">&nbsp;</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">Sincerely, </p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">MohammadHadi
  Najafi</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">Phone Number:
  +98-9179527931</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">Skype ID: m.h.najafi75<p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">&nbsp;</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif"></p>
  </td>"""


htmtext="""<td width="708" valign="top" style="width:531pt;border:1pt solid windowtext;padding:0in 5.4pt">
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">Dear
  Professor Kachroo,</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">I hope this
  email finds you well. My name is MohammadHadi Najafi, and I am writing to
  express my strong interest in pursuing a Direct PhD under your supervision in
  the Department of Electrical and Computer Engineering
  at University of Nevada, Las Vegas.
  I am particularly drawn to your groundbreaking work in Intelligent transportation systems,
  and I believe my background aligns well with your research interests.</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">&nbsp;</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">I have
  completed my Bachelor’s degree in Electrical Engineering at Quchan University
  of Technology. Throughout my academic journey, I have developed a deep
  passion for robotics and embedded systems design. My Bachelor’s thesis
  focuses on “Software-Based TFT LCD Driver for ARM Microcontrollers without
  LTDC”, which has prepared me well for advanced development of embedded
  systems.</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">&nbsp;</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">What sets me
  apart is my track record of success in both academic competitions and
  industry experience:</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif"><b>1.</b> <b>Robotics
  Competitions</b>: I have achieved significant success in several national and
  international robotics competitions. </p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif"><b>2.</b> <b>Professional
  Experience</b>: I have 5 years of hands-on experience in electronics and
  embedded system design. </p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif"><b>3.</b> <b>Technical
  Skills</b>: I possess strong skills in C, Python, MATLAB and LabView
  programming, Embedded Linux, PCB design, IOT and sensor integration. </p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">&nbsp;</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">I believe my
  combination of theoretical knowledge, practical experience, and
  problem-solving skills would allow me to contribute significantly to this
  line of research and expand upon it. I would be grateful for the opportunity
  to discuss potential PhD positions in your lab and how my background might
  complement your current research initiatives. I have attached my CV, which
  provides more details about my projects and achievements. If you need any
  additional information or work samples, please don't hesitate to ask.</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">&nbsp;</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">Thank you for
  your time and consideration. I am excited about the possibility of
  contributing to the cutting-edge research in your lab and look forward to the
  opportunity to discuss this further.</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">&nbsp;</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">Sincerely, </p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">MohammadHadi
  Najafi</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">Phone Number:
  +98-9179527931</p>
  <p class="MsoNormal" style="margin:0in;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif">Skype ID: m.h.najafi75</p>
  </td>"""




wb = openpyxl.load_workbook('DeadLine.xlsx')
ws = wb.active
ws = wb['6.0']

Start=1
End  =20
# LastNames=      [ws.cell(row=i,column=5).value for i in range(Start+1,End+2)]
# Departments=    [ws.cell(row=i,column=2).value for i in range(Start+1,End+2)]
# Universities=   [ws.cell(row=i,column=3).value for i in range(Start+1,End+2)]
# Interests=      [ws.cell(row=i,column=6).value for i in range(Start+1,End+2)]
# Emails=      [ws.cell(row=i,column=6).value for i in range(Start+1,End+2)]


# get the Gmail API service
service = gmail_authenticate()


# Email="mhnch75@gmail.com"
# LastName="Najafi"
# Department="Department of Electrical and Computer Engineering"
# University="University of Harward"
# Interest="Medical MEMS"


Start=1828
End=1844



for index in range(Start,End+1):

    i=index+1
    University=     ws.cell(row=i,column=2).value
    Department=     ws.cell(row=i,column=3).value
    LastName=       ws.cell(row=i,column=5).value
    Interest=       ws.cell(row=i,column=6).value
    ttry=           ws.cell(row=i,column=8).value
    Response=       ws.cell(row=i,column=11).value
    Title=          ws.cell(row=i,column=12).value

    Email=          ws.cell(row=i,column=9).value
    # Email="mhnch75@gmail.com"


    if not ttry:
        ttry=1
    else:
        ttry+=1


    print(Title)

    Message=  text1 + LastName + text2 + Department + text3 + University + text4 + Interest + text5
    Report=str(index) + "\t: " + str(ttry) + "\t: " +  University + "\t-\t" + LastName
    print(Report)

    if Response:
        if Response=="Auto" or Response=="Not now":
            Title="Inquiry about PhD opportunities under your supervision for fall 2025"
        else:
            Title=0
    else:
        if Title:
            pass
        else:
            Title="Inquiry about PhD opportunities under your supervision for fall 2025"


    if Title:
        try:
            result=send_message(service, Email , Title, 
                        Message, ["M.H.Najafi_July2024.pdf"])

            print(result['labelIds'][0])
            print("-------------------------------------------------------------------------------")

            ws.cell(row=i,column=8).value=ttry
            wb.save("DeadLine.xlsx")

            time.sleep(5)
            
            
        except:
            print("Error")
            print("-------------------------------------------------------------------------------")
    else:
        print("Already responsed")
        print("-------------------------------------------------------------------------------")






